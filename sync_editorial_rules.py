# -*- coding: utf-8 -*-
"""
Convert editorial_rules.xlsx -> editorial_rules.txt
Runs automatically before spellcheck; skips if xlsx is older than txt.
"""
import os
import sys


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    xlsx_path = os.path.join(script_dir, "editorial_rules.xlsx")
    txt_path = os.path.join(script_dir, "editorial_rules.txt")

    if not os.path.exists(xlsx_path):
        return

    # Only convert if xlsx is newer than txt
    if os.path.exists(txt_path):
        if os.path.getmtime(xlsx_path) <= os.path.getmtime(txt_path):
            return

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("openpyxl not installed, skipping xlsx sync", file=sys.stderr)
        return

    # Clean up old error sentinel
    err_path = txt_path + ".sync_error"
    try:
        if os.path.exists(err_path):
            os.remove(err_path)
    except OSError:
        pass

    try:
        wb = load_workbook(xlsx_path, read_only=True)
    except Exception as e:
        print(f"Cannot read xlsx (file locked?): {e}", file=sys.stderr)
        # Write sentinel so JSX knows sync failed
        try:
            with open(err_path, 'w', encoding='utf-8') as ef:
                ef.write(str(e))
        except OSError:
            pass
        return

    try:
        ws = wb.active

        lines = [
            "# \u0410\u0432\u0442\u043E\u0433\u0435\u043D\u0435\u0440\u0430\u0446\u0438\u044F \u0438\u0437 editorial_rules.xlsx\n",
            "# \u041D\u0435 \u0440\u0435\u0434\u0430\u043A\u0442\u0438\u0440\u0443\u0439\u0442\u0435 \u0432\u0440\u0443\u0447\u043D\u0443\u044E \u2014 \u0440\u0435\u0434\u0430\u043A\u0442\u0438\u0440\u0443\u0439\u0442\u0435 xlsx \u0444\u0430\u0439\u043B\n",
            "#\n",
            "# \u0424\u043E\u0440\u043C\u0430\u0442\u044B (Tab-separated):\n",
            "#   GREP<Tab>\u043F\u0430\u0442\u0442\u0435\u0440\u043D<Tab>\u0437\u0430\u043C\u0435\u043D\u0430[<Tab>\u0441\u0442\u0438\u043B\u044C \u0430\u0431\u0437\u0430\u0446\u0430]\n",
            "#   \u043D\u0430\u0439\u0442\u0438<Tab>\u0437\u0430\u043C\u0435\u043D\u0438\u0442\u044C[<Tab>\u0441\u0442\u0438\u043B\u044C \u0430\u0431\u0437\u0430\u0446\u0430]\n",
        ]

        def _cell(row, idx):
            if idx >= len(row) or row[idx] is None:
                return ""
            return str(row[idx]).strip()

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 2:
                continue

            col_a = _cell(row, 0)   # Type
            col_b = _cell(row, 1)   # Find
            col_c = _cell(row, 2)   # Replace
            col_d = _cell(row, 3)   # Comment (not written as field, only as # line)
            col_e = _cell(row, 4)   # Paragraph style (scope) — optional

            if not col_b:
                continue

            rule_type = col_a.upper()

            # Add comment line above rule if present
            if col_d:
                lines.append(f"# {col_d}\n")

            # Scope field gets appended as trailing tab-separated column when non-empty.
            # Empty scope → rule applies globally (no trailing tab).
            scope_suffix = f"\t{col_e}" if col_e else ""

            if rule_type == "GREP" and col_c:
                lines.append(f"GREP\t{col_b}\t{col_c}{scope_suffix}\n")
            elif rule_type in ("TEXT", "\u0422\u0415\u041A\u0421\u0422", "") and col_b and col_c:
                lines.append(f"{col_b}\t{col_c}{scope_suffix}\n")
    finally:
        wb.close()

    try:
        with open(txt_path, 'w', encoding='utf-8') as f:
            f.writelines(lines)
    except OSError as e:
        print(f"Cannot write {txt_path}: {e}", file=sys.stderr)
        return

    print(f"Synced {len(lines) - 2} rules from xlsx", file=sys.stderr)


if __name__ == '__main__':
    main()
